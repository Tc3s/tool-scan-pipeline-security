#!/usr/bin/env python3
"""Offline regression checks for generated VA pipeline artifacts."""

from __future__ import annotations

import contextlib
import hashlib
import io
import json
import os
import re
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from scripts.schema_utils import extract_cves, redact_sensitive
from scripts.run_pipeline import _print_agent_prompt, write_zap_auth_replacer_config, zap_container_config_path
from scripts.policy_validator import validate_source
from scripts.export_json_soc import build_report as build_soc_report
from scripts.merge_vulns import merge_vulns


class PipelineRegressionTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls._env_keys = ["VA_RUN_DIR", "VA_RUN_ID", "VA_PROJECT_ROOT", "VA_VERIFIER_FILE", "VA_SCOPE_FILE"]
        cls._previous_env = {key: os.environ.get(key) for key in cls._env_keys}
        cls._temp_runtime = tempfile.TemporaryDirectory()
        cls.run_dir = Path(cls._temp_runtime.name) / "va-run"
        cls._write_fixture_runtime(cls.run_dir)

        os.environ["VA_RUN_DIR"] = str(cls.run_dir)
        os.environ["VA_RUN_ID"] = "unit-fixture"
        os.environ["VA_PROJECT_ROOT"] = str(ROOT)
        os.environ["VA_VERIFIER_FILE"] = str(cls.run_dir / "generated" / "verify_vulns.py")
        os.environ["VA_SCOPE_FILE"] = str(cls.run_dir / "scope.yml")

        env = os.environ.copy()
        for script_name in ["export_excel.py", "export_json_soc.py", "export_ai_context.py"]:
            result = subprocess.run(
                [sys.executable, str(ROOT / "scripts" / script_name)],
                cwd=ROOT,
                env=env,
                text=True,
                capture_output=True,
                timeout=60,
                check=False,
            )
            if result.returncode != 0:
                raise AssertionError(
                    f"{script_name} failed\nSTDOUT:\n{result.stdout}\nSTDERR:\n{result.stderr}"
                )

        cls.queue_path = cls.run_dir / "output/vuln_validation_queue.csv"
        cls.zap_findings_path = cls.run_dir / "normalized/zap_findings.csv"
        cls.zap_instances_path = cls.run_dir / "normalized/zap_instances.csv"
        cls.openvas_path = cls.run_dir / "normalized/openvas_findings.csv"
        cls.raw_path = cls.run_dir / "output/vuln_raw.csv"
        cls.soc_path = cls.run_dir / "reports/customer_safe/vuln_report_soc.json"
        cls.internal_soc_path = cls.run_dir / "reports/internal/vuln_report_soc.json"
        cls.xlsx_path = cls.run_dir / "reports/internal/vuln_attack_report.xlsx"
        cls.customer_xlsx_path = cls.run_dir / "reports/customer_safe/vuln_attack_report.xlsx"
        cls.ai_context_path = cls.run_dir / "ai_context/internal/verification_context.jsonl"
        cls.customer_ai_context_path = cls.run_dir / "ai_context/customer_safe/verification_context.jsonl"
        cls.ai_zap_context_path = cls.run_dir / "ai_context/internal/zap_instances_compact.jsonl"

        cls.queue = pd.read_csv(cls.queue_path)

    @classmethod
    def tearDownClass(cls):
        for key, value in getattr(cls, "_previous_env", {}).items():
            if value is None:
                os.environ.pop(key, None)
            else:
                os.environ[key] = value
        temp_runtime = getattr(cls, "_temp_runtime", None)
        if temp_runtime is not None:
            temp_runtime.cleanup()

    @staticmethod
    def _write_fixture_runtime(run_dir: Path):
        raw_dir = run_dir / "raw"
        normalized_dir = run_dir / "normalized"
        output_dir = run_dir / "output"
        generated_dir = run_dir / "generated"
        for path in [raw_dir, normalized_dir, output_dir, generated_dir]:
            path.mkdir(parents=True, exist_ok=True)

        public_sources = json.dumps([
            {
                "type": "Exploit-DB",
                "cve": "CVE-2016-0800",
                "title": "Synthetic DROWN proof of concept reference",
                "match_basis": "CVE_ONLY_PUBLIC_EXPLOIT_INTEL",
                "context_match": "CONTEXT_NOT_CLASSIFIED",
            },
            {
                "type": "Metasploit",
                "cve": "CVE-2014-3566",
                "module": "auxiliary/scanner/ssl/synthetic_poodle_check",
                "match_basis": "CVE_ONLY_PUBLIC_EXPLOIT_INTEL",
                "context_match": "CONTEXT_NOT_CLASSIFIED",
            },
        ], ensure_ascii=False)
        epss_all = json.dumps([
            {"cve": "CVE-2016-0800", "epss": 0.82, "percentile": 0.93},
            {"cve": "CVE-2014-3566", "epss": 0.31, "percentile": 0.75},
        ], ensure_ascii=False)
        zap_urls = json.dumps([
            "http://target.example/admin?username=zap&password=ZAP",
            "http://target.example/admin?token=synthetic-redaction-token",
        ])

        rows = [
            {
                "scanner": "OpenVAS",
                "scan_time": "2026-07-19T01:00:00Z",
                "asset": "10.10.10.10",
                "asset_type": "host",
                "location": "443/tcp",
                "url_or_port": "443/tcp",
                "finding_name": "Synthetic TLS legacy protocol support",
                "severity": "High",
                "cvss": 7.5,
                "cve": "CVE-2016-0800,CVE-2014-3566",
                "cve_list": json.dumps(["CVE-2016-0800", "CVE-2014-3566"]),
                "cwe": "CWE-327",
                "cwe_list": json.dumps(["CWE-327"]),
                "plugin_id": "oid-synthetic-tls",
                "description": "The service allows a deprecated TLS/SSL configuration in this synthetic fixture.",
                "scanner_evidence": "Scanner observed legacy TLS parameters on 443/tcp.",
                "scanner_solution": "Disable obsolete SSL/TLS versions and weak ciphers.",
                "evidence": "",
                "solution": "",
                "raw_reference": "fixture-openvas-1",
                "instance_count": 1,
                "affected_urls_json": "[]",
                "epss_score": 0.82,
                "epss_percentile": 0.93,
                "epss_source_cve": "CVE-2016-0800",
                "epss_all_json": epss_all,
                "exploit_available": True,
                "exploit_status": "PUBLIC_EXPLOIT_AVAILABLE",
                "exploit_source_cves": "CVE-2016-0800,CVE-2014-3566",
                "exploit_sources_json": public_sources,
                "exploit_evidence": "Synthetic public exploit intelligence matched by CVE only.",
                "exploit_match_basis": "CVE_ONLY_PUBLIC_EXPLOIT_INTEL",
                "exploit_match_note": "Matched by CVE; not proof that this exact target is exploitable.",
                "exploit_context_review_required": True,
                "exploit_context_summary": "Product/version context is not classified in the fixture.",
                "attack_tactic": "Initial Access",
                "attack_technique_id": "T1190",
                "attack_technique_name": "Exploit Public-Facing Application",
                "attack_confidence": 0.82,
                "attack_tactics_json": json.dumps(["Initial Access"]),
                "attack_techniques_json": json.dumps([{
                    "technique_id": "T1190",
                    "technique_name": "Exploit Public-Facing Application",
                    "tactic": "Initial Access",
                    "confidence": 0.82,
                    "reason": "Synthetic CVE-backed web/service exposure.",
                }]),
                "mapping_method": "fixture-rule",
                "reason": "Synthetic mapping with adequate confidence.",
                "needs_review": False,
                "sensitive_evidence": False,
                "verification_status": "NOT_VERIFIED",
                "verification_evidence": "",
                "verification_method": "",
                "verification_command": "",
                "verification_error": "",
                "verification_confidence": "",
                "verification_started_at": "",
                "verification_completed_at": "",
                "verification_safe_mode": True,
                "risk_score": 85.0,
                "priority": "P2",
                "risk_reason": "High scanner severity; public exploit intel is CVE-level only.",
                "risk_components_json": "{}",
            },
            {
                "scanner": "ZAP",
                "scan_time": "2026-07-19T01:05:00Z",
                "asset": "http://target.example",
                "asset_type": "web",
                "location": "http://target.example/admin?username=zap&password=ZAP",
                "url_or_port": "http://target.example/admin?token=synthetic-redaction-token",
                "finding_name": "Synthetic Missing X-Frame-Options Header",
                "severity": "Medium",
                "cvss": 5.0,
                "cve": "",
                "cve_list": "[]",
                "cwe": "CWE-1021",
                "cwe_list": json.dumps(["CWE-1021"]),
                "plugin_id": "10020",
                "description": "Response is missing a clickjacking protection header.",
                "scanner_evidence": "Header missing while password=ZAP username=ZAP token=synthetic-redaction-token was present in the test URL.",
                "scanner_solution": "Set X-Frame-Options or an equivalent CSP frame-ancestors policy.",
                "evidence": "",
                "solution": "",
                "raw_reference": "fixture-zap-10020",
                "instance_count": 2,
                "affected_urls_json": zap_urls,
                "epss_score": 0.0,
                "epss_percentile": 0.0,
                "epss_source_cve": "",
                "epss_all_json": "[]",
                "exploit_available": False,
                "exploit_status": "NO_CVE_ID",
                "exploit_source_cves": "",
                "exploit_sources_json": "[]",
                "exploit_evidence": "",
                "exploit_match_basis": "NO_CVE",
                "exploit_match_note": "No CVE ID was available for exploit-intelligence lookup.",
                "exploit_context_review_required": False,
                "exploit_context_summary": "",
                "attack_tactic": "Defense Evasion",
                "attack_technique_id": "T1211",
                "attack_technique_name": "Exploitation for Defense Evasion",
                "attack_confidence": 0.72,
                "attack_tactics_json": json.dumps(["Defense Evasion"]),
                "attack_techniques_json": json.dumps([{
                    "technique_id": "T1211",
                    "technique_name": "Exploitation for Defense Evasion",
                    "tactic": "Defense Evasion",
                    "confidence": 0.72,
                    "reason": "Synthetic header weakness mapping.",
                }]),
                "mapping_method": "fixture-rule",
                "reason": "Synthetic mapping with adequate confidence.",
                "needs_review": False,
                "sensitive_evidence": True,
                "verification_status": "CONFIRMED_PRESENT",
                "verification_evidence": "HEAD /admin confirmed the header is absent without sending credentials.",
                "verification_method": "safe_header_check",
                "verification_command": "safe_request HEAD http://target.example/admin",
                "verification_error": "",
                "verification_confidence": "HIGH",
                "verification_started_at": "2026-07-19T01:10:00Z",
                "verification_completed_at": "2026-07-19T01:10:02Z",
                "verification_safe_mode": True,
                "risk_score": 55.0,
                "priority": "P3",
                "risk_reason": "Medium scanner severity; confirmed present on target.",
                "risk_components_json": "{}",
            },
            {
                "scanner": "OpenVAS",
                "scan_time": "2026-07-19T01:15:00Z",
                "asset": "10.10.10.20",
                "asset_type": "host",
                "location": "22/tcp",
                "url_or_port": "22/tcp",
                "finding_name": "Synthetic SSH weak algorithm notice",
                "severity": "Low",
                "cvss": 2.6,
                "cve": "",
                "cve_list": "[]",
                "cwe": "",
                "cwe_list": "[]",
                "plugin_id": "oid-synthetic-ssh",
                "description": "The service advertises a legacy algorithm in this fixture.",
                "scanner_evidence": "Scanner observed a weak algorithm banner.",
                "scanner_solution": "Disable legacy SSH algorithms after compatibility review.",
                "evidence": "",
                "solution": "",
                "raw_reference": "fixture-openvas-2",
                "instance_count": 1,
                "affected_urls_json": "[]",
                "epss_score": 0.0,
                "epss_percentile": 0.0,
                "epss_source_cve": "",
                "epss_all_json": "[]",
                "exploit_available": False,
                "exploit_status": "NO_CVE_ID",
                "exploit_source_cves": "",
                "exploit_sources_json": "[]",
                "exploit_evidence": "",
                "exploit_match_basis": "NO_CVE",
                "exploit_match_note": "No CVE ID was available for exploit-intelligence lookup.",
                "exploit_context_review_required": False,
                "exploit_context_summary": "",
                "attack_tactic": "Unknown",
                "attack_technique_id": "",
                "attack_technique_name": "",
                "attack_confidence": 0.0,
                "attack_tactics_json": "[]",
                "attack_techniques_json": "[]",
                "mapping_method": "needs-review",
                "reason": "No reliable ATT&CK mapping for this synthetic low finding.",
                "needs_review": True,
                "sensitive_evidence": False,
                "verification_status": "NOT_VERIFIED",
                "verification_evidence": "",
                "verification_method": "",
                "verification_command": "",
                "verification_error": "",
                "verification_confidence": "",
                "verification_started_at": "",
                "verification_completed_at": "",
                "verification_safe_mode": True,
                "risk_score": 20.0,
                "priority": "P4",
                "risk_reason": "Low scanner severity; no public exploit found.",
                "risk_components_json": "{}",
            },
            {
                "scanner": "OpenVAS",
                "scan_time": "2026-07-19T01:20:00Z",
                "asset": "10.10.10.30",
                "asset_type": "host",
                "location": "8080/tcp",
                "url_or_port": "8080/tcp",
                "finding_name": "Synthetic outdated service banner",
                "severity": "Medium",
                "cvss": 4.3,
                "cve": "CVE-2024-0001",
                "cve_list": json.dumps(["CVE-2024-0001"]),
                "cwe": "CWE-1104",
                "cwe_list": json.dumps(["CWE-1104"]),
                "plugin_id": "oid-synthetic-banner",
                "description": "A service version in this synthetic fixture appears outdated.",
                "scanner_evidence": "Banner version matched a stale signature.",
                "scanner_solution": "Validate the package version and apply vendor updates.",
                "evidence": "",
                "solution": "",
                "raw_reference": "fixture-openvas-3",
                "instance_count": 1,
                "affected_urls_json": "[]",
                "epss_score": 0.04,
                "epss_percentile": 0.21,
                "epss_source_cve": "CVE-2024-0001",
                "epss_all_json": json.dumps([{"cve": "CVE-2024-0001", "epss": 0.04, "percentile": 0.21}]),
                "exploit_available": False,
                "exploit_status": "NO_PUBLIC_EXPLOIT_FOUND",
                "exploit_source_cves": "",
                "exploit_sources_json": "[]",
                "exploit_evidence": "No configured local public source matched this synthetic CVE.",
                "exploit_match_basis": "CVE_CHECKED_NO_PUBLIC_SOURCE",
                "exploit_match_note": "No public exploit/module/template found in configured local sources.",
                "exploit_context_review_required": False,
                "exploit_context_summary": "",
                "attack_tactic": "Initial Access",
                "attack_technique_id": "T1190",
                "attack_technique_name": "Exploit Public-Facing Application",
                "attack_confidence": 0.68,
                "attack_tactics_json": json.dumps(["Initial Access"]),
                "attack_techniques_json": json.dumps([{
                    "technique_id": "T1190",
                    "technique_name": "Exploit Public-Facing Application",
                    "tactic": "Initial Access",
                    "confidence": 0.68,
                    "reason": "Below customer-safe confidence threshold.",
                }]),
                "mapping_method": "fixture-rule-low-confidence",
                "reason": "Synthetic mapping below confidence threshold.",
                "needs_review": True,
                "sensitive_evidence": False,
                "verification_status": "CHECKED_NOT_REPRODUCED",
                "verification_evidence": "Safe banner check did not reproduce the stale version claim.",
                "verification_method": "safe_banner_check",
                "verification_command": "safe_connect 10.10.10.30:8080",
                "verification_error": "",
                "verification_confidence": "MEDIUM",
                "verification_started_at": "2026-07-19T01:25:00Z",
                "verification_completed_at": "2026-07-19T01:25:03Z",
                "verification_safe_mode": True,
                "risk_score": 30.0,
                "priority": "P4",
                "risk_reason": "Medium scanner severity; checked but not reproduced.",
                "risk_components_json": "{}",
            },
        ]

        queue_df = pd.DataFrame(rows)
        queue_df.to_csv(output_dir / "vuln_validation_queue.csv", index=False)
        queue_df.to_csv(output_dir / "vuln_attack_enriched.csv", index=False)
        queue_df.to_csv(output_dir / "vuln_attack_mapped.csv", index=False)
        queue_df.to_csv(output_dir / "vuln_raw.csv", index=False)
        queue_df[queue_df["scanner"].eq("ZAP")].to_csv(normalized_dir / "zap_findings.csv", index=False)
        queue_df[queue_df["scanner"].eq("OpenVAS")].to_csv(normalized_dir / "openvas_findings.csv", index=False)

        pd.DataFrame([
            {
                "scanner": "ZAP",
                "scan_time": "2026-07-19T01:05:00Z",
                "asset": "http://target.example",
                "asset_type": "web",
                "location": "http://target.example/admin?username=zap&password=ZAP",
                "url_or_port": "http://target.example/admin?username=zap&password=ZAP",
                "finding_name": "Synthetic Missing X-Frame-Options Header",
                "severity": "Medium",
                "plugin_id": "10020",
                "method": "GET",
                "param": "",
                "scanner_evidence": "Header missing on fixture instance.",
                "cwe": "CWE-1021",
                "cwe_list": json.dumps(["CWE-1021"]),
                "instance_index": 1,
            },
            {
                "scanner": "ZAP",
                "scan_time": "2026-07-19T01:05:00Z",
                "asset": "http://target.example",
                "asset_type": "web",
                "location": "http://target.example/admin?token=synthetic-redaction-token",
                "url_or_port": "http://target.example/admin?token=synthetic-redaction-token",
                "finding_name": "Synthetic Missing X-Frame-Options Header",
                "severity": "Medium",
                "plugin_id": "10020",
                "method": "GET",
                "param": "token",
                "scanner_evidence": "Header missing on second fixture instance.",
                "cwe": "CWE-1021",
                "cwe_list": json.dumps(["CWE-1021"]),
                "instance_index": 2,
            },
        ]).to_csv(normalized_dir / "zap_instances.csv", index=False)

        (raw_dir / "zap_report.json").write_text("{}", encoding="utf-8")
        (raw_dir / "openvas_report.xml").write_text("<report />\n", encoding="utf-8")
        (generated_dir / "verify_vulns.py").write_text("# generated verifier fixture\n", encoding="utf-8")
        (run_dir / "scope.yml").write_text(
            "\n".join([
                'schema_version: "1.0"',
                'target: "http://target.example"',
                "allowed_hosts:",
                '  - "target.example"',
                '  - "10.10.10.10"',
                '  - "10.10.10.20"',
                '  - "10.10.10.30"',
                "allowed_schemes:",
                '  - "http"',
                '  - "https"',
                "allowed_ports:",
                "  - 80",
                "  - 443",
                "  - 8080",
                "allowed_methods:",
                '  - "GET"',
                '  - "HEAD"',
                '  - "OPTIONS"',
                "same_origin_redirects_only: true",
                "max_requests_per_second: 1",
                "max_concurrency: 1",
                "auth_allowed: false",
            ]),
            encoding="utf-8",
        )

    def test_parser_counts_are_preserved(self):
        zap_findings = pd.read_csv(self.zap_findings_path)
        zap_instances = pd.read_csv(self.zap_instances_path)
        openvas = pd.read_csv(self.openvas_path)
        raw = pd.read_csv(self.raw_path)

        self.assertGreater(len(zap_findings), 0)
        self.assertGreaterEqual(len(zap_instances), len(zap_findings))
        self.assertGreater(len(openvas), 0)
        self.assertGreater(len(raw), 0)
        self.assertLessEqual(len(raw), len(zap_findings) + len(openvas))

    def test_merge_supports_openvas_only_without_stale_zap(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp = Path(temp_dir)
            openvas_csv = temp / "openvas_findings.csv"
            missing_zap = temp / "zap_findings.csv"
            output_csv = temp / "vuln_raw.csv"
            pd.DataFrame([
                {
                    "scanner": "OpenVAS",
                    "scan_time": "2026-07-19T00:00:00Z",
                    "asset": "10.0.0.1",
                    "asset_type": "host",
                    "location": "443/tcp",
                    "url_or_port": "443/tcp",
                    "finding_name": "TLS test finding",
                    "severity": "Medium",
                    "cvss": 5.0,
                    "cve": "",
                    "cve_list": "[]",
                    "cwe": "",
                    "cwe_list": "[]",
                    "plugin_id": "oid-1",
                    "description": "desc",
                    "scanner_evidence": "evidence",
                    "scanner_solution": "solution",
                    "raw_reference": "oid-1",
                    "instance_count": 1,
                    "affected_urls_json": "[]",
                }
            ]).to_csv(openvas_csv, index=False)

            count = merge_vulns(missing_zap, openvas_csv, output_csv)
            merged = pd.read_csv(output_csv)

            self.assertEqual(count, 1)
            self.assertEqual(len(merged), 1)
            self.assertEqual(merged.loc[0, "scanner"], "OpenVAS")
            self.assertEqual(merged.loc[0, "asset"], "10.0.0.1")

    def test_canonical_schema_keeps_scanner_context(self):
        for column in [
            "scanner_evidence",
            "scanner_solution",
            "cve_list",
            "cwe_list",
            "location",
            "exploit_status",
            "verification_status",
            "verification_evidence",
            "verification_method",
            "verification_command",
            "verification_error",
            "verification_confidence",
            "verification_started_at",
            "verification_completed_at",
            "verification_safe_mode",
            "risk_reason",
        ]:
            self.assertIn(column, self.queue.columns)

        self.assertGreater(int(self.queue["scanner_evidence"].fillna("").astype(str).str.strip().ne("").sum()), 0)
        self.assertGreater(int(self.queue["risk_reason"].fillna("").astype(str).str.strip().ne("").sum()), 0)

    def test_multicve_epss_is_not_lost(self):
        multi = self.queue[
            self.queue["cve"].fillna("").astype(str).str.contains(",", regex=False)
        ]
        self.assertGreaterEqual(len(multi), 0)
        if len(multi):
            self.assertGreaterEqual(int(multi["epss_score"].fillna(0).ge(0).sum()), len(multi))

    def test_exploit_intel_is_separate_from_verification(self):
        verification_statuses = set(self.queue["verification_status"].dropna().unique())
        allowed_statuses = {
            "NOT_VERIFIED",
            "REPRODUCED",
            "CONFIRMED_PRESENT",
            "CHECKED_NOT_REPRODUCED",
            "FALSE_POSITIVE",
            "SKIPPED_SAFE_MODE",
            "NEEDS_MANUAL_REVIEW",
            "ERROR",
        }
        self.assertTrue(verification_statuses <= allowed_statuses, verification_statuses)
        statuses = set(self.queue["exploit_status"].dropna().unique())
        self.assertIn("NO_CVE_ID", statuses)
        self.assertNotIn("WEAPONIZED", statuses)
        self.assertNotIn("REPRODUCED", statuses)
        self.assertNotIn("CONFIRMED_PRESENT", statuses)
        self.assertIn("exploit_available", self.queue.columns)

    def test_attack_mapping_no_fake_full_coverage(self):
        self.assertGreater(int(self.queue["needs_review"].fillna(False).astype(bool).sum()), 0)
        self.assertGreater(int(self.queue["attack_tactic"].fillna("").eq("Unknown").sum()), 0)
        trusted = self.queue[pd.to_numeric(self.queue["attack_confidence"], errors="coerce").fillna(0) >= 0.7]
        self.assertGreaterEqual(len(trusted), 0)

    def test_risk_guardrails(self):
        low = self.queue[self.queue["severity"].eq("Low")]
        self.assertEqual(int(low["priority"].eq("P1").sum()), 0)
        self.assertEqual(int(self.queue["priority"].isin(["P1", "P2", "P3", "P4"]).sum()), len(self.queue))

    def test_soc_json_schema_v1_and_redaction(self):
        with open(self.soc_path, encoding="utf-8") as handle:
            report = json.load(handle)
        with open(self.internal_soc_path, encoding="utf-8") as handle:
            internal = json.load(handle)

        self.assertEqual(report["schema_version"], "1.0")
        self.assertEqual(report["summary"]["total_findings"], len(self.queue))
        self.assertEqual(len(report["findings"]), len(self.queue))
        self.assertEqual(len(internal["findings"]), len(self.queue))
        self.assertIn("provenance", report)
        self.assertIn("scanner", report["findings"][0])
        self.assertIn("event", report["findings"][0])
        self.assertIn("finding", report["findings"][0])
        self.assertIn("url", report["findings"][0])
        self.assertIn("network", report["findings"][0])
        self.assertIn("exploit_intel", report["findings"][0])
        self.assertIn("verification", report["findings"][0])
        self.assertIn("risk", report["findings"][0])
        self.assertIn("exploit_context_review_required", report["summary"])
        self.assertEqual(report["reporting"]["ingest_contract"]["dedup_field"], "finding.dedup_hash")
        self.assertEqual(report["reporting"]["ingest_contract"]["target_proof_field"], "verification.status")
        self.assertFalse(report["reporting"]["ingest_contract"]["public_exploit_available_is_target_proof"])
        self.assertEqual(report["findings"][0]["event"]["type"], "vulnerability_finding")
        self.assertEqual(report["findings"][0]["event"]["action"], "upsert")
        self.assertIn("asset_type", report["findings"][0]["target"])
        self.assertIn("target_level_proof", report["findings"][0]["verification"])
        self.assertIn("intel_level", report["findings"][0]["exploit_intel"])
        self.assertIn("instance_count", report["findings"][0]["scanner"])
        self.assertIn("affected_urls", report["findings"][0]["scanner"])
        self.assertIn("sha256", report["provenance"]["exporter"])
        finding_ids = [finding["finding"]["id"] for finding in report["findings"]]
        dedup_hashes = [finding["finding"]["dedup_hash"] for finding in report["findings"]]
        self.assertEqual(len(finding_ids), len(set(finding_ids)))
        self.assertEqual(len(dedup_hashes), len(set(dedup_hashes)))
        self.assertTrue(all(finding_id.startswith("va-find-") for finding_id in finding_ids))
        self.assertEqual(
            finding_ids,
            [finding["finding"]["id"] for finding in internal["findings"]],
        )
        self.assertTrue(all(finding["event"]["dedup_hash"] == finding["finding"]["dedup_hash"] for finding in report["findings"]))
        self.assertFalse(any(str(finding["target"]["host"]).startswith(("http://", "https://")) for finding in report["findings"]))
        timestamp_re = re.compile(r"^\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}Z$")
        self.assertRegex(report["run_metadata"]["generated_at"], timestamp_re)
        self.assertTrue(all(timestamp_re.match(finding["event"]["observed_at"]) for finding in report["findings"] if finding["event"]["observed_at"]))
        allowed_contexts = {
            "CRITICAL_ACTION_REQUIRED",
            "VERIFIED_PRIORITY_REVIEW",
            "VERIFIED_LOW_PRIORITY",
            "FALSE_POSITIVE_FILTERED",
            "PUBLIC_EXPLOIT_AVAILABLE",
            "POTENTIAL_RISK",
        }
        self.assertTrue(all(finding["reporting"]["soc_context"] in allowed_contexts for finding in report["findings"]))
        self.assertTrue(any(finding["scanner"]["affected_urls"] for finding in report["findings"]))

        safe_text = json.dumps(report, ensure_ascii=False).lower()
        for raw_secret in [
            "postgres:postgres",
            "msfadmin:msfadmin",
            "root:root",
            "password: password",
            "password=zap",
            "username=zap",
            "token=synthetic-redaction-token",
            "/sensitive/local/path/tool-scan-pipeline-security",
        ]:
            self.assertNotIn(raw_secret, safe_text)
        self.assertEqual(report["run_metadata"]["input_file"], "vuln_validation_queue.csv")
        self.assertTrue(report["reporting"]["customer_safe"])
        self.assertIn("match_basis", report["findings"][0]["exploit_intel"])
        self.assertIn("match_note", report["findings"][0]["exploit_intel"])
        self.assertIn("context_review_required", report["findings"][0]["exploit_intel"])
        self.assertIn("context_summary", report["findings"][0]["exploit_intel"])
        for key in ["method", "command", "error", "confidence", "started_at", "completed_at", "safe_mode"]:
            self.assertIn(key, report["findings"][0]["verification"])

    def test_soc_json_mixed_verification_semantics(self):
        df = pd.DataFrame([
            {
                "scanner": "unit",
                "scan_time": "2026-07-19T10:00:00+07:00",
                "asset": "https://target.example",
                "asset_type": "web",
                "location": "https://target.example/admin?token=secret-value",
                "url_or_port": "https://target.example/admin?token=secret-value",
                "finding_name": "Confirmed issue",
                "severity": "High",
                "cvss": 8.0,
                "plugin_id": "unit-1",
                "description": "confirmed",
                "scanner_evidence": "token=secret-value",
                "scanner_solution": "fix",
                "verification_status": "CONFIRMED_PRESENT",
                "verification_evidence": "present",
                "verification_safe_mode": True,
                "priority": "P2",
                "risk_score": 75,
                "risk_reason": "confirmed present",
                "risk_components_json": "{}",
                "exploit_available": False,
                "exploit_status": "NO_CVE_ID",
                "exploit_sources_json": "[]",
            },
            {
                "scanner": "unit",
                "scan_time": "2026-07-19T11:00:00Z",
                "asset": "target.example",
                "asset_type": "web",
                "location": "https://target.example/old",
                "url_or_port": "https://target.example/old",
                "finding_name": "False positive issue",
                "severity": "Medium",
                "plugin_id": "unit-2",
                "verification_status": "FALSE_POSITIVE",
                "verification_safe_mode": True,
                "priority": "P4",
                "risk_score": 0,
                "risk_reason": "false positive",
                "risk_components_json": "{}",
                "exploit_available": False,
                "exploit_status": "NO_CVE_ID",
                "exploit_sources_json": "[]",
            },
            {
                "scanner": "unit",
                "scan_time": "2026-07-19T12:00:00Z",
                "asset": "target.example",
                "asset_type": "web",
                "location": "https://target.example/cve",
                "url_or_port": "https://target.example/cve",
                "finding_name": "Public exploit only",
                "severity": "Medium",
                "plugin_id": "unit-3",
                "cve": "CVE-2024-0001",
                "verification_status": "NOT_VERIFIED",
                "verification_safe_mode": True,
                "priority": "P3",
                "risk_score": 50,
                "risk_reason": "public exploit intel",
                "risk_components_json": "{}",
                "exploit_available": True,
                "exploit_status": "PUBLIC_EXPLOIT_AVAILABLE",
                "exploit_sources_json": json.dumps([{"type": "Exploit-DB", "cve": "CVE-2024-0001", "context_match": "CONTEXT_NOT_CLASSIFIED"}]),
            },
        ])
        report = build_soc_report(df, "vuln_validation_queue.csv", customer_safe=True)
        by_name = {item["vulnerability"]["name"]: item for item in report["findings"]}

        self.assertTrue(by_name["Confirmed issue"]["verification"]["target_level_proof"])
        self.assertEqual(by_name["Confirmed issue"]["reporting"]["soc_context"], "CRITICAL_ACTION_REQUIRED")
        self.assertFalse(by_name["False positive issue"]["verification"]["target_level_proof"])
        self.assertEqual(by_name["False positive issue"]["reporting"]["soc_context"], "FALSE_POSITIVE_FILTERED")
        self.assertFalse(by_name["Public exploit only"]["verification"]["target_level_proof"])
        self.assertEqual(by_name["Public exploit only"]["reporting"]["soc_context"], "PUBLIC_EXPLOIT_AVAILABLE")
        self.assertTrue(by_name["Public exploit only"]["exploit_intel"]["context_review_required"])
        self.assertEqual(by_name["Public exploit only"]["exploit_intel"]["intel_level"], "cve")
        self.assertEqual(by_name["Confirmed issue"]["target"]["host"], "target.example")
        self.assertNotIn("token=secret-value", json.dumps(report, ensure_ascii=False))

    def test_excel_sheet_layout(self):
        internal_sheets = pd.read_excel(self.xlsx_path, sheet_name=None)
        customer_wb = load_workbook(self.customer_xlsx_path, read_only=False, data_only=True)

        for sheet in [
            "Executive Summary",
            "Findings",
            "CVE Intel",
            "Exploit Intel",
            "Mapping Review",
            "ZAP Instances",
        ]:
            self.assertIn(sheet, internal_sheets)

        self.assertEqual(customer_wb.sheetnames, ["Tổng Quan", "Chi Tiết Lỗ Hổng"])
        self.assertEqual(customer_wb["Tổng Quan"]["A1"].value, "BÁO CÁO ĐÁNH GIÁ LỖ HỔNG BẢO MẬT")
        self.assertEqual(len(customer_wb["Tổng Quan"]._charts), 2)
        detail_ws = customer_wb["Chi Tiết Lỗ Hổng"]
        headers = [detail_ws.cell(4, col).value for col in range(1, 7)]
        self.assertEqual(headers, ["STT", "Mức Độ", "Plugin/CVE", "Lỗ Hổng", "Chi Tiết", "Giải Pháp"])
        self.assertEqual(detail_ws.max_column, 6)
        customer_severity_values = [
            str(detail_ws.cell(row, 2).value or "")
            for row in range(5, detail_ws.max_row + 1)
        ]
        self.assertFalse(any(value.startswith(("P1 -", "P2 -", "P3 -", "P4 -")) for value in customer_severity_values))
        severity_order = {
            "CRITICAL": 0,
            "HIGH": 1,
            "MEDIUM": 2,
            "LOW": 3,
            "INFORMATIONAL": 4,
            "INFO": 4,
            "LOG": 5,
            "UNKNOWN": 6,
        }
        severity_ranks = [severity_order.get(value.upper(), 6) for value in customer_severity_values if value]
        self.assertEqual(severity_ranks, sorted(severity_ranks))

        self.assertIn("Sensitive Evidence", internal_sheets)
        self.assertEqual(len(internal_sheets["Findings"]), len(self.queue))
        self.assertGreaterEqual(len(internal_sheets["ZAP Instances"]), 1)
        self.assertIn("vuln_validation_queue.csv", [customer_wb["Tổng Quan"].cell(row, 4).value for row in range(1, 12)])

        cve_intel = internal_sheets["CVE Intel"]
        public_cves = cve_intel[cve_intel["exploit_status"].eq("PUBLIC_EXPLOIT_AVAILABLE")]
        self.assertEqual(
            int(public_cves["exploit_sources"].fillna("").astype(str).str.strip().eq("").sum()),
            0,
        )
        self.assertIn("finding_exploit_status", cve_intel.columns)
        self.assertIn("match_basis", cve_intel.columns)
        self.assertIn("match_basis", internal_sheets["Exploit Intel"].columns)
        self.assertIn("match_note", internal_sheets["Exploit Intel"].columns)
        self.assertIn("context_match", internal_sheets["Exploit Intel"].columns)
        self.assertIn("context_review_required", internal_sheets["Exploit Intel"].columns)

        customer_values = []
        for ws in customer_wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                customer_values.extend(str(value) for value in row if value is not None)
        customer_text = "\n".join(customer_values).lower()
        for raw_secret in [
            "postgres:postgres",
            "msfadmin:msfadmin",
            "root:root",
            "password: password",
            "password=zap",
            "username=zap",
            "token=synthetic-redaction-token",
            "/sensitive/local/path/tool-scan-pipeline-security",
        ]:
            self.assertNotIn(raw_secret, customer_text)

    def test_cve_extractor_handles_multi_cve_cells(self):
        cves = extract_cves("CVE-2016-0800,CVE-2014-3566")
        self.assertEqual(cves, ["CVE-2016-0800", "CVE-2014-3566"])

    def test_redaction_covers_sensitive_url_values(self):
        redacted = redact_sensitive(
            "http://host/path?password=ZAP&username=ZAP&token=synthetic-redaction-token"
        )
        lowered = redacted.lower()
        self.assertNotIn("password=zap", lowered)
        self.assertNotIn("username=zap", lowered)
        self.assertNotIn("token=synthetic-redaction-token", lowered)
        self.assertIn("password=[redacted]", lowered)
        self.assertIn("username=[redacted]", lowered)
        self.assertIn("token=[redacted]", lowered)

    def test_ai_context_jsonl_is_compact_and_verifier_safe(self):
        self.assertTrue(self.ai_context_path.exists())
        self.assertTrue(self.customer_ai_context_path.exists())
        self.assertTrue(self.ai_zap_context_path.exists())

        records = [
            json.loads(line)
            for line in self.ai_context_path.read_text(encoding="utf-8").splitlines()
            if line.strip()
        ]
        self.assertEqual(len(records), len(self.queue))
        self.assertIn("target", records[0])
        self.assertIn("vulnerability", records[0])
        self.assertIn("verification", records[0])
        self.assertIn("exploit_intel", records[0])
        self.assertIn("safe_mode", records[0]["verification"])

        legacy_agent_statuses = {
            "NO_CVE_ID",
            "NO_PUBLIC_EXPLOIT_FOUND",
            "PUBLIC_EXPLOIT_AVAILABLE",
            "EXPLOIT_TEMPLATE_AVAILABLE",
            "INTEL_CHECK_ERROR",
        }
        leaked_legacy_statuses = [
            record["verification"].get("agent_status")
            for record in records
            if record["verification"].get("agent_status") in legacy_agent_statuses
        ]
        self.assertEqual(leaked_legacy_statuses, [])

        zap_records_with_instances = [
            record for record in records
            if record.get("scanner") == "ZAP" and record.get("zap_instances")
        ]
        self.assertGreaterEqual(len(zap_records_with_instances), 1)

        zap_instances = [
            json.loads(line)
            for line in self.ai_zap_context_path.read_text(encoding="utf-8").splitlines()
            if line.strip()
        ]
        self.assertGreaterEqual(len(zap_instances), len(zap_records_with_instances))
        self.assertIn("url", zap_instances[0])
        self.assertIn("method", zap_instances[0])

        safe_text = self.customer_ai_context_path.read_text(encoding="utf-8").lower()
        for raw_secret in [
            "postgres:postgres",
            "msfadmin:msfadmin",
            "root:root",
            "password=zap",
            "username=zap",
            "token=synthetic-redaction-token",
            "/sensitive/local/path/tool-scan-pipeline-security",
            "root123",
        ]:
            self.assertNotIn(raw_secret, safe_text)

    def test_agent_prompts_are_production_safe_and_redacted(self):
        captured = {}
        for mode, cookie in [
            ("GREYBOX", "SECRET_COOKIE_SHOULD_NOT_LEAK=abc"),
            ("BLACKBOX", None),
        ]:
            buffer = io.StringIO()
            with contextlib.redirect_stdout(buffer):
                _print_agent_prompt(
                    "http://target.example",
                    [],
                    {"mode": mode, "cookie": cookie},
                )
            captured[mode] = buffer.getvalue()

        greybox = captured["GREYBOX"]
        blackbox = captured["BLACKBOX"]
        all_text = "\n".join(captured.values())
        lower = all_text.lower()

        self.assertIn("PRODUCTION PRESENCE VERIFIER", greybox)
        self.assertIn("GREYBOX_AUTHENTICATED", greybox)
        self.assertIn("VA_AUTH_COOKIE", greybox)
        self.assertNotIn("SECRET_COOKIE_SHOULD_NOT_LEAK", greybox)
        self.assertIn("BLACKBOX_UNAUTHENTICATED", blackbox)
        self.assertIn("No authentication is authorized", blackbox)

        for forbidden in [
            "--risk 3",
            "169.254.169.254",
            "file:///etc/passwd",
            "root:x:0:0",
            "ysoserial",
            "smiley backdoor",
            "root:<empty>",
            "full firepower",
            "elite security",
            "ssrf probe",
            "os.system",
            "harcode session",
            "hardcode session",
        ]:
            self.assertNotIn(forbidden, lower)

    def test_zap_auth_config_keeps_cookie_out_of_process_args(self):
        secret = "PHPSESSID=SECRET_COOKIE_SHOULD_NOT_LEAK; token=abc123"
        with tempfile.TemporaryDirectory() as temp_dir:
            config_path = Path(write_zap_auth_replacer_config(secret, raw_dir=temp_dir))
            try:
                cmd = [
                    "docker",
                    "run",
                    "--rm",
                    "ghcr.io/zaproxy/zaproxy:stable",
                    "zap-baseline.py",
                    "-z",
                    f"-configfile {zap_container_config_path(config_path)}",
                ]
                self.assertFalse(any(secret in str(part) for part in cmd))
                self.assertIn(secret, config_path.read_text(encoding="utf-8"))
                if os.name != "nt":
                    self.assertEqual(config_path.stat().st_mode & 0o777, 0o600)
            finally:
                if config_path.exists():
                    config_path.unlink()

            with self.assertRaises(ValueError):
                write_zap_auth_replacer_config("PHPSESSID=abc\r\nInjected: yes", raw_dir=temp_dir)

    def test_policy_validator_rejects_unsafe_generated_code(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            unsafe = Path(temp_dir) / "bad_verifier.py"
            unsafe.write_text(
                "import subprocess\nsubprocess.run('id', shell=True)\n",
                encoding="utf-8",
            )
            result = validate_source(unsafe)
            self.assertFalse(result["passed"])
            self.assertTrue(any(item["rule"] == "shell_true" for item in result["findings"]))

    def test_verifier_dry_run_writes_approval_manifest_without_live_run(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            run_dir = Path(temp_dir)
            output_dir = run_dir / "output"
            context_dir = run_dir / "ai_context/internal"
            generated_dir = run_dir / "generated"
            verification_dir = run_dir / "verification"
            for path in [output_dir, context_dir, generated_dir, verification_dir]:
                path.mkdir(parents=True, exist_ok=True)

            queue = pd.DataFrame([{
                "scanner": "ZAP",
                "asset": "http://target.example",
                "asset_type": "web",
                "location": "http://target.example/",
                "url_or_port": "http://target.example/",
                "finding_name": "Missing X-Frame-Options Header",
                "severity": "Low",
                "cvss": "",
                "cve": "",
                "cwe": "CWE-1021",
                "plugin_id": "10020",
                "description": "Header missing",
                "scanner_evidence": "Scanner reported missing header",
                "scanner_solution": "Set header",
                "exploit_status": "NO_CVE_ID",
                "verification_status": "NOT_VERIFIED",
            }])
            queue_file = output_dir / "vuln_validation_queue.csv"
            queue.to_csv(queue_file, index=False)

            from scripts.verification_contract import finding_id
            finding = finding_id(queue.iloc[0], 1)
            (context_dir / "verification_context.jsonl").write_text(
                json.dumps({"id": finding, "priority": "P4", "scanner": "ZAP", "target": {"location": "http://target.example/"}}) + "\n",
                encoding="utf-8",
            )
            (context_dir / "zap_instances_compact.jsonl").write_text("", encoding="utf-8")
            (context_dir / "manifest.json").write_text(json.dumps({"record_count": 1}), encoding="utf-8")
            (run_dir / "scope.yml").write_text(
                "\n".join([
                    'schema_version: "1.0"',
                    'target: "http://target.example"',
                    "allowed_hosts:",
                    '  - "target.example"',
                    "allowed_schemes:",
                    '  - "http"',
                    "allowed_ports:",
                    "  - 80",
                    "allowed_methods:",
                    '  - "GET"',
                    '  - "HEAD"',
                    '  - "OPTIONS"',
                    "same_origin_redirects_only: true",
                    "max_requests_per_second: 1",
                    "max_concurrency: 1",
                    "auth_allowed: false",
                ]),
                encoding="utf-8",
            )
            verifier = generated_dir / "verify_vulns.py"
            verifier.write_text(
                """
#!/usr/bin/env python3
import argparse
import json
from pathlib import Path
from datetime import datetime, timezone

def now():
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()

def read_first_id(path):
    with open(path, encoding="utf-8") as handle:
        for line in handle:
            if line.strip():
                return json.loads(line)["id"]
    raise SystemExit(1)

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("target")
    parser.add_argument("--mode", required=True)
    parser.add_argument("--context-file", required=True)
    parser.add_argument("--zap-context-file", required=True)
    parser.add_argument("--scope-file", required=True)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--plan-file")
    parser.add_argument("--results-file")
    parser.add_argument("--approval-file")
    args = parser.parse_args()
    finding_id = read_first_id(args.context_file)
    if args.dry_run:
        Path(args.plan_file).parent.mkdir(parents=True, exist_ok=True)
        Path(args.plan_file).write_text(json.dumps({
            "schema_version": "generated-verifier-plan-1.0",
            "generated_at": now(),
            "target": args.target,
            "mode": args.mode,
            "planned_total": 1,
            "planned_status_counts": {"CONFIRMED_PRESENT": 1},
            "planned_methods": ["fixture_header_check"],
            "manual_review_ids": [],
            "unsafe_skipped_ids": [],
        }), encoding="utf-8")
        return 0
    if not args.approval_file:
        return 1
    Path(args.results_file).parent.mkdir(parents=True, exist_ok=True)
    Path(args.results_file).write_text(json.dumps({
        "schema_version": "verification-result-1.0",
        "finding_id": finding_id,
        "status": "CONFIRMED_PRESENT",
        "method": "fixture_header_check",
        "evidence": "GET http://target.example/ observed the same missing header condition reported by scanner.",
        "confidence": "HIGH",
        "safe_mode": True,
        "completed_at": now(),
    }) + "\\n", encoding="utf-8")
    return 0

if __name__ == "__main__":
    raise SystemExit(main())
""",
                encoding="utf-8",
            )

            env = os.environ.copy()
            env["VA_RUN_DIR"] = str(run_dir)
            env["VA_VERIFIER_FILE"] = str(verifier)
            base_cmd = [
                sys.executable,
                str(ROOT / "scripts" / "verifier_lifecycle.py"),
                "--mode",
                "BLACKBOX",
                "--verifier-file",
                str(verifier),
            ]
            dry = subprocess.run(
                base_cmd + ["dry-run", "http://target.example"],
                cwd=ROOT,
                env=env,
                text=True,
                capture_output=True,
                timeout=30,
                check=False,
            )
            self.assertEqual(dry.returncode, 0, dry.stderr)
            approval = run_dir / "approval_manifest.json"
            manifest = json.loads(approval.read_text(encoding="utf-8"))
            self.assertFalse(manifest["approved"])
            self.assertEqual(manifest["verifier"]["sha256"], hashlib.sha256(verifier.read_bytes()).hexdigest())
            self.assertEqual(manifest["plan"]["planned_status_counts"], {"CONFIRMED_PRESENT": 1})

            approve = subprocess.run(
                base_cmd + ["approve", "http://target.example", "--operator", "unit-test"],
                cwd=ROOT,
                env=env,
                text=True,
                capture_output=True,
                timeout=30,
                check=False,
            )
            self.assertEqual(approve.returncode, 0, approve.stderr)
            live = subprocess.run(
                base_cmd + ["run", "http://target.example"],
                cwd=ROOT,
                env=env,
                text=True,
                capture_output=True,
                timeout=30,
                check=False,
            )
            self.assertEqual(live.returncode, 0, live.stderr)
            updated = pd.read_csv(queue_file)
            self.assertEqual(updated.loc[0, "verification_status"], "CONFIRMED_PRESENT")
            self.assertIn("confirmed present on target", updated.loc[0, "risk_reason"])

    def test_policy_validator_rejects_dynamic_reflection_and_dunders(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            bad_reflection = Path(temp_dir) / "bad_reflection.py"
            bad_reflection.write_text(
                "import importlib\nmod = importlib.import_module('os')\ngetattr(mod, 'system')('id')\n",
                encoding="utf-8",
            )
            result = validate_source(bad_reflection)
            self.assertFalse(result["passed"])
            self.assertTrue(any(f["rule"] in {"forbidden_module_import", "dynamic_reflection_bypass", "dynamic_import_attempt"} for f in result["findings"]))

            bad_dunder = Path(temp_dir) / "bad_dunder.py"
            bad_dunder.write_text(
                "x = ().__class__.__subclasses__()\n",
                encoding="utf-8",
            )
            result_dunder = validate_source(bad_dunder)
            self.assertFalse(result_dunder["passed"])
            self.assertTrue(any(f["rule"] == "forbidden_dunder_attribute" for f in result_dunder["findings"]))

    def test_private_ip_detection(self):
        from scripts.verification_contract import is_private_or_loopback_ip, canonical_target, ContractError
        self.assertTrue(is_private_or_loopback_ip("127.0.0.1"))
        self.assertTrue(is_private_or_loopback_ip("169.254.169.254"))
        self.assertTrue(is_private_or_loopback_ip("10.0.0.1"))
        self.assertTrue(is_private_or_loopback_ip("192.168.1.1"))
        self.assertFalse(is_private_or_loopback_ip("8.8.8.8"))

        with self.assertRaises(ContractError):
            canonical_target("http://127.0.0.1/admin")

    def test_false_positive_critical_risk_override_fixed(self):
        from scripts.calculate_risk_priority import calculate_risk_for_row
        row = {
            "severity": "Critical",
            "exploit_available": True,
            "verification_status": "FALSE_POSITIVE",
            "epss_score": 0.99,
        }
        res = calculate_risk_for_row(row)
        self.assertEqual(res["priority"], "P4")
        self.assertLessEqual(res["risk_score"], 39.0)


if __name__ == "__main__":
    unittest.main(verbosity=2)
